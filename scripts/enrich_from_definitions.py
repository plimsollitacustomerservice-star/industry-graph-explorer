"""
enrich_from_definitions.py
==========================
Reads the master XLS (Industry-Definitions-Database-Enhanced_with_full_matrix.xls)
from the repo root, extracts ATECO codes, keywords, and adjacent industries
from the definition text columns, and regenerates the data/ JSON files.

Run from the repo root:
    python scripts/enrich_from_definitions.py

Requirements:
    pip install openpyxl xlrd pandas
"""

import re, json, os, glob
from collections import defaultdict

# ── Try xlrd for .xls, openpyxl for .xlsx ─────────────────────────────────
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Locate master XLS/XLSX ─────────────────────────────────────────────────
SEARCH_PATHS = [
    '.',           # repo root
    'data',        # data/ folder
]

def find_master_file():
    for folder in SEARCH_PATHS:
        for pattern in ['*.xlsx', '*.xls']:
            matches = glob.glob(os.path.join(folder, pattern))
            # Prefer the enhanced/full matrix file
            for m in matches:
                if 'enhanced' in m.lower() or 'matrix' in m.lower() or 'definition' in m.lower():
                    return m
            if matches:
                return matches[0]
    return None

master_path = find_master_file()
if not master_path:
    raise FileNotFoundError("No XLS/XLSX master file found in repo root or data/")
print(f"Using master file: {master_path}")

# ── Load all rows into a list of dicts ─────────────────────────────────────
def load_xls_xlrd(path):
    book = xlrd.open_workbook(path)
    # Use the sheet with the most rows
    ws = sorted(book.sheets(), key=lambda s: s.nrows, reverse=True)[0]
    print(f"Sheet: '{ws.name}' | rows: {ws.nrows} | cols: {ws.ncols}")
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    print(f"Headers: {headers[:20]}")
    rows = []
    for r in range(1, ws.nrows):
        row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
        rows.append(row)
    return headers, rows

def load_xlsx_openpyxl(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Use the sheet with the most rows
    best = max(wb.worksheets, key=lambda s: s.max_row or 0)
    print(f"Sheet: '{best.title}' | rows: {best.max_row}")
    headers = []
    rows = []
    for i, row in enumerate(best.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(v).strip() if v is not None else '' for v in row]
            print(f"Headers: {headers[:20]}")
        else:
            rows.append({headers[c]: row[c] for c in range(len(headers))})
    return headers, rows

if master_path.endswith('.xls') and HAS_XLRD:
    headers, raw_rows = load_xls_xlrd(master_path)
elif HAS_OPENPYXL:
    headers, raw_rows = load_xlsx_openpyxl(master_path)
else:
    raise ImportError("Install xlrd (for .xls) or openpyxl (for .xlsx): pip install xlrd openpyxl")

print(f"Loaded {len(raw_rows)} raw rows")

# ── Fuzzy column finder ─────────────────────────────────────────────────────
col_lower = {h.lower(): h for h in headers}

def find_header(*candidates):
    for c in candidates:
        if c in headers: return c
        cl = c.lower()
        if cl in col_lower: return col_lower[cl]
        for h in col_lower:
            if cl in h: return col_lower[h]
    return None

H_RC   = find_header('RepCode', 'Rep Code', 'Code', 'ID')
H_NEN  = find_header('NameEnglish', 'Name English', 'Title', 'NameEN')
H_NIT  = find_header('NameNative', 'Name Native', 'NameIT', 'Italian Name')
H_SEC  = find_header('Sector', 'SectorCode')
H_ATP  = find_header('ATECOPrimary', 'ATECO Primary', 'Ateco', 'ATECO', 'NACE')
H_ATAL = find_header('ATECOAll', 'ATECO All', 'AllATECO')
H_PRI  = find_header('Priority', 'priority')
H_RDEN = find_header('ReportDefinitionEN', 'Report Definition EN', 'DefinitionEN', 'Definition EN', 'ReportDef', 'Description')
H_MDEN = find_header('MarketingDefinitionEN', 'Marketing Definition EN', 'MarketingDef')
H_KWEN = find_header('KeywordsIncludeEN', 'Keywords EN', 'KeywordsEN', 'Keywords Include EN', 'Keywords')
H_KWIT = find_header('KeywordsIncludeIT', 'Keywords IT', 'KeywordsIT')
H_ORB  = find_header('OrbisBoolean', 'Orbis Boolean', 'Orbis', 'Boolean')
H_TRA  = find_header('TradeAssociations', 'Trade Associations', 'Associations')
H_ADJ  = find_header('AdjacentIndustries', 'Adjacent Industries', 'Adjacent', 'Related Industries')
H_VCS  = find_header('ValueChainStage', 'Value Chain Stage', 'ValueChain', 'Stage')

print("\nColumn mapping:")
for n, h in [('RepCode', H_RC), ('NameEnglish', H_NEN), ('NameNative', H_NIT),
             ('Sector', H_SEC), ('ATECOPrimary', H_ATP), ('ATECOAll', H_ATAL),
             ('Priority', H_PRI), ('ReportDefinitionEN', H_RDEN),
             ('MarketingDefinitionEN', H_MDEN), ('KeywordsIncludeEN', H_KWEN),
             ('KeywordsIncludeIT', H_KWIT), ('OrbisBoolean', H_ORB),
             ('TradeAssociations', H_TRA), ('AdjacentIndustries', H_ADJ),
             ('ValueChainStage', H_VCS)]:
    print(f"  {n:30s} -> {repr(h)}")

# ── Helper: safe string from cell value ────────────────────────────────────
def sv(row, key, maxlen=0):
    if key is None: return ''
    v = row.get(key, '')
    if v is None: return ''
    s = str(v).strip()
    if s in ('None', 'nan', 'NULL', 'null', '', '<NA>'): return ''
    # Strip float suffix like "123.0"
    if s.endswith('.0') and s[:-2].isdigit(): s = s[:-2]
    return s[:maxlen] if maxlen else s

# ── ATECO extraction from free text ────────────────────────────────────────
# Matches patterns: 10.11, 10.11.1, 47, C10, C10.1
ATECO_RE = re.compile(
    r'\b([A-Z]?\d{1,2}(?:\.\d{1,2}(?:\.\d{1,2})?)?)\b'
)
ATECO_SECTION_RE = re.compile(r'\b([A-U])\b')  # single-letter NACE sections

def extract_ateco_from_text(text):
    """Extract ATECO/NACE codes from definition text."""
    if not text: return []
    codes = []
    for m in ATECO_RE.finditer(text):
        code = m.group(1)
        # Filter out obvious non-codes (years, phone numbers, etc.)
        if re.match(r'^\d{4}$', code): continue  # 4-digit year
        if len(code) == 1: continue               # single digit
        codes.append(code)
    return list(dict.fromkeys(codes))  # deduplicate, preserve order

# ── Keyword extraction from free text ──────────────────────────────────────
STOP_WORDS = {
    'the','and','or','of','in','to','for','a','an','is','are','that','this',
    'with','as','at','by','from','on','be','which','it','its','also','such',
    'their','they','these','those','can','may','will','has','have','been',
    'more','other','used','use','used','using','e','g','etc','i','e',
    'including','include','includes','manufacture','manufacturer','manufacturers',
    'production','producer','producers','product','products','supplier','suppliers',
    'company','companies','industry','industries','sector','business',
    'services','service','provide','providing'
}

def extract_keywords_from_text(text, name_en=''):
    """Extract meaningful keywords from definition text + industry name."""
    if not text and not name_en: return []
    combined = f"{name_en} {text}"
    # Extract capitalised phrases and nouns
    words = re.findall(r'\b[A-Za-z][a-z]{2,}\b', combined)
    keywords = []
    seen = set()
    for w in words:
        wl = w.lower()
        if wl in STOP_WORDS or len(wl) < 3: continue
        if wl not in seen:
            seen.add(wl)
            keywords.append(w.lower())
    # Also extract 2-word phrases from the name
    name_words = name_en.split()
    for i in range(len(name_words) - 1):
        phrase = f"{name_words[i]} {name_words[i+1]}".lower()
        if phrase not in seen and len(phrase) > 6:
            seen.add(phrase)
            keywords.insert(0, phrase)
    return keywords[:20]  # cap at 20 keywords

# ── Adjacent industry extraction from definition text ──────────────────────
# Looks for patterns like "related to", "upstream", "downstream", sector refs
ADJ_TRIGGER_RE = re.compile(
    r'(?:upstream|downstream|adjacent|related|supplied by|supplies to|'
    r'works with|linked to|connected to|alongside|in conjunction with|'
    r'sister industry|value chain|supply chain)',
    re.IGNORECASE
)

def extract_adjacent_from_text(text, rep_code, all_industries_by_name):
    """
    Try to find adjacent industry RepCodes by matching industry names
    mentioned in the definition text.
    Returns a comma-separated string of RepCodes.
    """
    if not text: return ''
    found = []
    text_lower = text.lower()
    for name, rc in all_industries_by_name.items():
        if rc == rep_code: continue
        if len(name) < 6: continue
        if name.lower() in text_lower:
            found.append(rc)
        if len(found) >= 8: break
    return ','.join(found[:8])

# ── Orbis Boolean builder from keywords ────────────────────────────────────
def build_orbis_boolean(name_en, keywords):
    """Build a simple Orbis/BvD Boolean search string from name + keywords."""
    parts = []
    # Core name as phrase
    clean_name = re.sub(r'\s+', ' ', name_en.strip())
    parts.append(f'"{clean_name}"')
    # Add top 5 keywords as OR alternatives
    kw_terms = [k for k in keywords[:10] if len(k.split()) <= 2 and len(k) > 4][:5]
    if kw_terms:
        parts.append('OR ' + ' OR '.join(f'"{k}"' for k in kw_terms))
    return ' '.join(parts)[:400]

# ── SECTOR inference fallback ───────────────────────────────────────────────
SECTOR_COLORS = {
    'MAN','WHL','RET','HEA','ICT','AGR','FIN','ENE','CON','EDU','SER'
}

def infer_sector(name):
    n = name.lower()
    if any(x in n for x in ['wholesale','ingrosso','distributor','distribution']): return 'WHL'
    if any(x in n for x in ['retail','shop','store','dealer']): return 'RET'
    if any(x in n for x in ['health','medical','pharma','hospital','clinic','dental']): return 'HEA'
    if any(x in n for x in ['software','it service','cloud','saas','tech','digital','cyber']): return 'ICT'
    if any(x in n for x in ['energy','solar','wind','electric power','oil','gas','nuclear','renewab']): return 'ENE'
    if any(x in n for x in ['construct','building','architect','civil engineer']): return 'CON'
    if any(x in n for x in ['agri','farm','crop','livestock','fishery','food production']): return 'AGR'
    if any(x in n for x in ['finance','bank','insurance','investment','fund']): return 'FIN'
    if any(x in n for x in ['education','school','university','training','e-learning']): return 'EDU'
    if any(x in n for x in ['manufactur','fabricat','production','producer','maker']): return 'MAN'
    return 'SER'

# ── Main parse loop ─────────────────────────────────────────────────────────
print("\nParsing industries...")
industries = []
seen_codes = set()

for row in raw_rows:
    rep_code = sv(row, H_RC)
    if not rep_code or rep_code in seen_codes: continue
    name_en = sv(row, H_NEN, 120)
    if not name_en: continue
    seen_codes.add(rep_code)

    # Sector
    sector = sv(row, H_SEC)
    if sector not in SECTOR_COLORS:
        sector = infer_sector(name_en)

    # Priority normalisation
    priority = sv(row, H_PRI)
    if priority and len(priority) > 10 and ('T' in priority or '-' in priority[:5]):
        priority = 'Medium'
    if priority not in ('High', 'Medium', 'Low'):
        priority = 'Medium'

    # Definition texts (these are the rich source)
    report_def = sv(row, H_RDEN, 800)
    mktg_def   = sv(row, H_MDEN, 400)
    full_text  = f"{report_def} {mktg_def}".strip()

    # ATECO: prefer explicit column, extract from text as fallback
    ateco_primary = sv(row, H_ATP, 15)
    ateco_all     = sv(row, H_ATAL, 300)
    if not ateco_primary or not ateco_all:
        extracted = extract_ateco_from_text(full_text)
        if extracted:
            if not ateco_primary:
                ateco_primary = extracted[0]
            if not ateco_all:
                ateco_all = ','.join(extracted[:10])

    # Keywords: prefer explicit column, extract from text as fallback
    kw_en = sv(row, H_KWEN, 300)
    kw_it = sv(row, H_KWIT, 300)
    extracted_keywords = []
    if not kw_en:
        extracted_keywords = extract_keywords_from_text(full_text, name_en)
        kw_en = ','.join(extracted_keywords[:15])

    # Orbis Boolean: prefer explicit column, build from keywords as fallback
    orbis = sv(row, H_ORB, 400)
    if not orbis and (name_en or extracted_keywords):
        orbis = build_orbis_boolean(name_en, extracted_keywords)

    industries.append({
        'RepCode':               rep_code,
        'NameEnglish':           name_en,
        'NameNative':            sv(row, H_NIT, 120),
        'Sector':                sector,
        'ValueChainStage':       sv(row, H_VCS, 50),
        'ATECOPrimary':          ateco_primary,
        'ATECOAll':              ateco_all,
        'Priority':              priority,
        'ReportDefinitionEN':    report_def,
        'MarketingDefinitionEN': mktg_def,
        'KeywordsIncludeEN':     kw_en,
        'KeywordsIncludeIT':     kw_it,
        'OrbisBoolean':          orbis,
        'TradeAssociations':     sv(row, H_TRA, 300),
        'AdjacentIndustries':    sv(row, H_ADJ, 300),  # will be enriched below
    })

print(f"Parsed {len(industries)} industries")

# ── Build name lookup for adjacency cross-referencing ──────────────────────
print("Building adjacency from definition cross-references...")
name_to_code = {}
for ind in industries:
    # Index by normalised name variants
    name_to_code[ind['NameEnglish'].lower()] = ind['RepCode']
    # Also index short form (first 3 words)
    short = ' '.join(ind['NameEnglish'].lower().split()[:3])
    if len(short) > 8:
        name_to_code[short] = ind['RepCode']

enriched = 0
for ind in industries:
    # Only fill if AdjacentIndustries is empty
    if ind['AdjacentIndustries']:
        continue
    full_text = f"{ind['ReportDefinitionEN']} {ind['MarketingDefinitionEN']}"
    adj = extract_adjacent_from_text(full_text, ind['RepCode'], name_to_code)
    if adj:
        ind['AdjacentIndustries'] = adj
        enriched += 1

print(f"Enriched {enriched} industries with adjacency from definition text")

# ── Sector-based peer adjacency fallback ───────────────────────────────────
# For industries still missing adjacency, assign top peers from same sector
# (by RepCode proximity — nearest numeric neighbors in same sector)
sector_map = defaultdict(list)
for ind in industries:
    sector_map[ind['Sector']].append(ind)

fallback_count = 0
for ind in industries:
    if ind['AdjacentIndustries']:
        continue
    peers = [p for p in sector_map[ind['Sector']] if p['RepCode'] != ind['RepCode']]
    if not peers:
        continue
    # Pick up to 5 peers with nearest RepCode numbers
    try:
        own_n = int(ind['RepCode'])
        peers_sorted = sorted(peers, key=lambda p: abs(int(p['RepCode']) - own_n))
        ind['AdjacentIndustries'] = ','.join(p['RepCode'] for p in peers_sorted[:5])
        fallback_count += 1
    except ValueError:
        pass

print(f"Applied sector-proximity fallback adjacency to {fallback_count} industries")

# ── Print sample ─────────────────────────────────────────────────────────────
print("\nSample enriched records:")
for ind in industries[:3]:
    print(f"  [{ind['RepCode']}] {ind['NameEnglish']}")
    print(f"    Sector:  {ind['Sector']}")
    print(f"    ATECO:   {ind['ATECOPrimary']} | All: {ind['ATECOAll'][:60]}")
    print(f"    Keywords:{ind['KeywordsIncludeEN'][:70]}")
    print(f"    Adjacent:{ind['AdjacentIndustries']}")
    print(f"    Orbis:   {ind['OrbisBoolean'][:80]}")
    print(f"    DefEN:   {ind['ReportDefinitionEN'][:100]}")
    print()

# ── Write sector JSON files ────────────────────────────────────────────────
print("Writing data/ JSON files...")
by_sector = defaultdict(list)
for ind in industries:
    by_sector[ind['Sector']].append(ind)

os.makedirs('data', exist_ok=True)
all_industry_files = []

for sector, items in sorted(by_sector.items()):
    # Dynamic chunk size: keep each JSON file under ~250 KB
    sample_size = len(json.dumps(items[:5], ensure_ascii=False))
    avg = sample_size / 5
    chunk_size = max(30, int(240_000 / avg))

    for i in range(0, len(items), chunk_size):
        n = i // chunk_size + 1
        fname = f"data/industries_{sector}_{n}.json" if len(items) > chunk_size else f"data/industries_{sector}.json"
        with open(fname, 'w', encoding='utf-8') as f:
            json.dump(items[i:i+chunk_size], f, ensure_ascii=False, separators=(',', ':'))
        all_industry_files.append(fname)
        sz = os.path.getsize(fname) // 1024
        print(f"  {fname}: {len(items[i:i+chunk_size])} items ({sz} KB)")

# ── Rebuild links from AdjacentIndustries field ───────────────────────────
print("Building links from AdjacentIndustries...")
links = []
link_set = set()
code_to_sector = {ind['RepCode']: ind['Sector'] for ind in industries}

for ind in industries:
    if not ind['AdjacentIndustries']:
        continue
    targets = [t.strip() for t in re.split(r'[,|;]', ind['AdjacentIndustries']) if t.strip()]
    for t in targets:
        if t not in code_to_sector:
            continue
        pair = tuple(sorted([ind['RepCode'], t]))
        if pair in link_set:
            continue
        link_set.add(pair)
        # Determine direction: same sector = Peer, else supply chain
        if code_to_sector.get(t) == ind['Sector']:
            direction = 'Peer'
        else:
            direction = 'Upstream' if int(ind['RepCode']) > int(t) else 'Downstream'
        links.append({
            'FromIndustryCode': ind['RepCode'],
            'ToIndustryCode':   t,
            'Direction':        direction,
            'StrengthScore':    3
        })

print(f"Built {len(links)} semantic links from adjacency data")

link_files = []
chunk = 3000
for i in range(0, min(len(links), 15000), chunk):
    n = i // chunk + 1
    fname = f'data/links_{n}.json'
    with open(fname, 'w', encoding='utf-8') as f:
        json.dump(links[i:i+chunk], f, ensure_ascii=False, separators=(',', ':'))
    link_files.append(fname)
    print(f"  {fname}: {len(links[i:i+chunk])} links")

# ── Write manifest ─────────────────────────────────────────────────────────
manifest = {
    'industryFiles':   sorted([f.replace('data/', '') for f in all_industry_files]),
    'linkFiles':       sorted([f.replace('data/', '') for f in link_files]),
    'totalIndustries': len(industries),
    'totalLinks':      min(len(links), 15000),
    'enrichmentNote':  'ATECO, keywords and adjacency extracted from definition text'
}
with open('data/manifest.json', 'w', encoding='utf-8') as f:
    json.dump(manifest, f, indent=2)

print(f"\nDone.")
print(f"  Industries : {len(industries)}")
print(f"  Links      : {min(len(links), 15000)}")
print(f"  Sector files: {len(all_industry_files)}")
print(f"\nNext: commit data/ folder and push to GitHub Pages.")
