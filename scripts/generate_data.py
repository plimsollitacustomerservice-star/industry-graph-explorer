import openpyxl, json, random, os, glob
from collections import defaultdict

print('Loading XLSX...')
xlsx_files = glob.glob('data/*.xlsx') + glob.glob('data/*.xls')
if not xlsx_files:
    raise FileNotFoundError('No XLSX/XLS file found in data/')
xlsx_path = sorted(xlsx_files)[-1]  # prefer .xlsx over .xls
print(f'Using: {xlsx_path}')

wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

# Find the sheet that has industry data (most rows)
sheet_names = wb.sheetnames
print(f'Sheets: {sheet_names}')
ws = wb.active
for sname in sheet_names:
    s = wb[sname]
    print(f'  Sheet "{sname}" - checking...')
    ws = s
    break

# Read header row and build column index map
headers = []
for row in ws.iter_rows(min_row=1, max_row=1):
    for i, cell in enumerate(row):
        v = str(cell.value).strip() if cell.value is not None else ''
        headers.append(v)
        print(f'  col[{i}] = {repr(v)}')

col = {h: i for i, h in enumerate(headers)}
print(f'Total columns: {len(headers)}')

# Column name fallbacks for fuzzy matching
def find_col(*candidates):
    for c in candidates:
        if c in col: return col[c]
        # case-insensitive
        cl = c.lower()
        for h in col:
            if h.lower() == cl: return col[h]
        # partial match
        for h in col:
            if cl in h.lower(): return col[h]
    return None

RC   = find_col('RepCode', 'Rep Code', 'rep_code', 'Code', 'ID', 'IndustryCode')
NEN  = find_col('NameEnglish', 'Name English', 'English Name', 'NameEN', 'Name_English', 'Title')
NIT  = find_col('NameNative', 'Name Native', 'NameIT', 'Italian Name', 'Name_Native', 'NomItaliano')
SEC  = find_col('Sector', 'SectorCode', 'Sector Code')
ATP  = find_col('ATECOPrimary', 'ATECO Primary', 'Ateco', 'ATECOCode', 'ATECO', 'NACE')
ATAL = find_col('ATECOAll', 'ATECO All', 'AllATECO', 'ATECO_All')
PRI  = find_col('Priority', 'priority')
RDEN = find_col('ReportDefinitionEN', 'Report Definition EN', 'ReportDefinition', 'Definition EN', 'DefinitionEN', 'ReportDef')
MDEN = find_col('MarketingDefinitionEN', 'Marketing Definition EN', 'MarketingDef', 'Marketing Definition')
KWEN = find_col('KeywordsIncludeEN', 'Keywords EN', 'KeywordsEN', 'Keywords Include EN', 'Keywords')
KWIT = find_col('KeywordsIncludeIT', 'Keywords IT', 'KeywordsIT', 'Keywords Include IT')
ORB  = find_col('OrbisBoolean', 'Orbis Boolean', 'Orbis', 'BooleanSearch', 'Boolean')
TRA  = find_col('TradeAssociations', 'Trade Associations', 'Associations')
ADJ  = find_col('AdjacentIndustries', 'Adjacent Industries', 'Adjacent', 'Related Industries')
VCS  = find_col('ValueChainStage', 'Value Chain Stage', 'ValueChain', 'Stage')

print(f'Column mapping:')
for name, idx in [('RepCode',RC),('NameEnglish',NEN),('NameNative',NIT),('Sector',SEC),
                   ('ATECOPrimary',ATP),('ATECOAll',ATAL),('Priority',PRI),
                   ('ReportDefinitionEN',RDEN),('MarketingDefinitionEN',MDEN),
                   ('KeywordsIncludeEN',KWEN),('KeywordsIncludeIT',KWIT),
                   ('OrbisBoolean',ORB),('TradeAssociations',TRA),
                   ('AdjacentIndustries',ADJ),('ValueChainStage',VCS)]:
    print(f'  {name} -> col[{idx}] = {repr(headers[idx]) if idx is not None else "NOT FOUND"}')

def cv(row, idx, maxlen=0):
    if idx is None: return ''
    try:
        v = row[idx].value
        if v is None: return ''
        s = str(v).strip()
        if s in ('None', 'nan', 'NULL', 'null', '<NA>'): return ''
        return s[:maxlen] if maxlen else s
    except: return ''

SECTOR_COLORS = {
    'MAN':'Manufacturing','WHL':'Wholesale','RET':'Retail','HEA':'Healthcare',
    'ICT':'ICT','AGR':'Agriculture','FIN':'Finance','ENE':'Energy',
    'CON':'Construction','EDU':'Education','SER':'Services'
}

def get_sector_from_name(name):
    n = str(name).lower()
    if 'wholesale' in n or 'ingrosso' in n: return 'WHL'
    if 'retail' in n or ' shop' in n or ' store' in n: return 'RET'
    if 'health' in n or 'medical' in n or 'hospital' in n or 'pharma' in n: return 'HEA'
    if 'software' in n or 'it service' in n or 'cloud' in n or ' tech ' in n: return 'ICT'
    if 'energy' in n or 'solar' in n or 'wind power' in n or 'electric power' in n: return 'ENE'
    if 'construct' in n or 'building' in n or 'architect' in n: return 'CON'
    if 'agri' in n or 'beverage' in n or 'farm' in n: return 'AGR'
    if 'finance' in n or 'banking' in n or 'insurance' in n: return 'FIN'
    if 'education' in n or 'school' in n or 'training' in n or 'university' in n: return 'EDU'
    if 'manufactur' in n or 'fabricat' in n: return 'MAN'
    return 'SER'

print('Parsing industries...')
industries = []
seen = set()
row_count = 0
for row in ws.iter_rows(min_row=2):
    if RC is None: break
    raw = row[RC].value
    if raw is None: continue
    row_count += 1
    rep_code = str(int(raw)) if isinstance(raw, (int,float)) and raw == int(raw) else str(raw).strip()
    if not rep_code or rep_code in seen or rep_code in ('None','nan',''): continue
    seen.add(rep_code)

    name_en = cv(row, NEN) if NEN is not None else ''
    if not name_en: continue

    # Sector: prefer explicit column, fall back to name inference
    sector = cv(row, SEC) if SEC is not None else ''
    if sector not in SECTOR_COLORS: sector = get_sector_from_name(name_en)

    ateco_raw = cv(row, ATP) if ATP is not None else ''
    ateco_all = cv(row, ATAL, 200) if ATAL is not None else ateco_raw[:200]
    if not ateco_all: ateco_all = ateco_raw

    # Clean ATECOPrimary to first code only
    ateco_primary = ateco_raw.split(',')[0].replace(' ','').strip()[:15] if ateco_raw else ''

    priority = cv(row, PRI) if PRI is not None else 'Medium'
    # Normalize priority (timestamps -> Medium)
    if priority and len(priority) > 10 and ('T' in priority or '-' in priority[:5]):
        priority = 'Medium'
    if priority not in ('High','Medium','Low'): priority = 'Medium'

    industries.append({
        'RepCode':               rep_code,
        'NameEnglish':           name_en[:120],
        'NameNative':            cv(row, NIT, 120),
        'Sector':                sector,
        'ValueChainStage':       cv(row, VCS, 50),
        'ATECOPrimary':          ateco_primary,
        'ATECOAll':              ateco_all,
        'Priority':              priority,
        'ReportDefinitionEN':    cv(row, RDEN, 400),
        'MarketingDefinitionEN': cv(row, MDEN, 280),
        'KeywordsIncludeEN':     cv(row, KWEN, 250),
        'KeywordsIncludeIT':     cv(row, KWIT, 250),
        'OrbisBoolean':          cv(row, ORB, 400),
        'TradeAssociations':     cv(row, TRA, 250),
        'AdjacentIndustries':    cv(row, ADJ, 200),
    })

print(f'Parsed {len(industries)} industries from {row_count} rows')

# Print sample
for i in industries[:3]:
    print(f'  Sample: {i["RepCode"]} | {i["NameEnglish"]} | Sector={i["Sector"]} | ATECO={i["ATECOPrimary"]} | Priority={i["Priority"]}')
    print(f'    ReportDef: {i["ReportDefinitionEN"][:80]}')
    print(f'    Keywords: {i["KeywordsIncludeEN"][:60]}')

by_sector = defaultdict(list)
for ind in industries:
    by_sector[ind['Sector']].append(ind)

os.makedirs('data', exist_ok=True)
all_industry_files = []

for sector, items in by_sector.items():
    sample_json = json.dumps(items[:5], ensure_ascii=False, separators=(',',':'))
    avg_size = len(sample_json) / 5
    chunk_size = max(50, int(240_000 / avg_size))

    if len(items) <= chunk_size:
        fname = f'data/industries_{sector}.json'
        with open(fname, 'w', encoding='utf-8') as f:
            json.dump(items, f, ensure_ascii=False, separators=(',', ':'))
        all_industry_files.append(fname)
        sz = os.path.getsize(fname) // 1024
        print(f'  {fname}: {len(items)} items ({sz} KB)')
    else:
        for i in range(0, len(items), chunk_size):
            n = i // chunk_size + 1
            fname = f'data/industries_{sector}_{n}.json'
            with open(fname, 'w', encoding='utf-8') as f:
                json.dump(items[i:i+chunk_size], f, ensure_ascii=False, separators=(',', ':'))
            all_industry_files.append(fname)
            sz = os.path.getsize(fname) // 1024
            print(f'  {fname}: {len(items[i:i+chunk_size])} items ({sz} KB)')

# Generate peer links within sectors
print('Generating links...')
random.seed(42)
links = []
link_set = set()
for ind in industries:
    same = [i for i in industries if i['Sector'] == ind['Sector'] and i['RepCode'] != ind['RepCode']]
    if len(same) < 2: continue
    for target in random.sample(same, min(5, len(same))):
        pair = tuple(sorted([ind['RepCode'], target['RepCode']]))
        if pair in link_set: continue
        link_set.add(pair)
        links.append({
            'FromIndustryCode': ind['RepCode'],
            'ToIndustryCode':   target['RepCode'],
            'Direction':        'Peer',
            'StrengthScore':    random.randint(2, 5)
        })

print(f'Generated {len(links)} links')

link_files = []
chunk = 3000
for i in range(0, min(len(links), 15000), chunk):
    n = i // chunk + 1
    fname = f'data/links_{n}.json'
    with open(fname, 'w', encoding='utf-8') as f:
        json.dump(links[i:i+chunk], f, ensure_ascii=False, separators=(',', ':'))
    link_files.append(fname)
    print(f'  {fname}: {len(links[i:i+chunk])} links')

manifest = {
    'industryFiles': sorted([f.replace('data/', '') for f in all_industry_files]),
    'linkFiles':     sorted([f.replace('data/', '') for f in link_files]),
    'totalIndustries': len(industries),
    'totalLinks':      min(len(links), 15000)
}
with open('data/manifest.json', 'w') as f:
    json.dump(manifest, f, indent=2)

print(f'Done. {len(industries)} industries, {min(len(links),15000)} links, {len(all_industry_files)} sector files.')
